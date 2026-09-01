from pathlib import Path

from openpyxl import load_workbook

from report_items.clients.items_client import Item
from report_items.reports.excel_report import ExcelReportGenerator
from report_items.reports.item_report_mapper import ItemReportMapper


def build_item(item_id: int, descripcion: str) -> Item:
    return Item.model_validate(
        {
            "idCompania": 1,
            "item": item_id,
            "descripcionItem": descripcion,
            "criterios": [],
            "descripcionesTecnicas": [],
        },
    )


class TestGenerate:
    def test_genera_el_archivo_con_headers_y_filas(self, tmp_path: Path):
        mapper = ItemReportMapper()
        generator = ExcelReportGenerator(mapper=mapper)

        items = [
            build_item(1, "Primer item"),
            build_item(2, "Segundo item"),
        ]

        output_path = generator.generate(
            items=items,
            output_directory=tmp_path,
        )

        assert output_path.exists()

        assert output_path.name == "items.xlsx"

        workbook = load_workbook(output_path)
        worksheet = workbook["Items"]

        rows = list(worksheet.iter_rows(values_only=True))

        assert rows[0] == tuple(ItemReportMapper.HEADERS)
        assert rows[1][ItemReportMapper.HEADERS.index("DESCRIP. ITEM")] == "Primer item"
        assert (
            rows[2][ItemReportMapper.HEADERS.index("DESCRIP. ITEM")] == "Segundo item"
        )

    def test_headers_se_escriben_en_negrita(self, tmp_path: Path):
        mapper = ItemReportMapper()
        generator = ExcelReportGenerator(mapper=mapper)

        output_path = generator.generate(
            items=[],
            output_directory=tmp_path,
        )

        workbook = load_workbook(output_path)
        worksheet = workbook["Items"]

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1))

        assert all(cell.font.bold for cell in header_row)

    def test_crea_el_directorio_de_salida_si_no_existe(self, tmp_path: Path):
        mapper = ItemReportMapper()
        generator = ExcelReportGenerator(mapper=mapper)

        output_directory = tmp_path / "reportes" / "anidados"

        output_path = generator.generate(
            items=[],
            output_directory=output_directory,
        )

        assert output_path.exists()

    def test_no_deja_archivo_temporal_tras_generar(self, tmp_path: Path):
        mapper = ItemReportMapper()
        generator = ExcelReportGenerator(mapper=mapper)

        generator.generate(items=[], output_directory=tmp_path)

        temp_files = list(tmp_path.glob("*.tmp.xlsx"))
        assert temp_files == []

    def test_ejecuciones_sucesivas_sobreescriben_el_mismo_archivo(
        self,
        tmp_path: Path,
    ):
        mapper = ItemReportMapper()
        generator = ExcelReportGenerator(mapper=mapper)

        first_run_path = generator.generate(
            items=[build_item(1, "Primer item")],
            output_directory=tmp_path,
        )

        second_run_path = generator.generate(
            items=[build_item(2, "Segundo item")],
            output_directory=tmp_path,
        )

        assert first_run_path == second_run_path

        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert xlsx_files == [second_run_path]

        workbook = load_workbook(second_run_path)
        worksheet = workbook["Items"]

        rows = list(worksheet.iter_rows(values_only=True))
        descripciones = [
            row[ItemReportMapper.HEADERS.index("DESCRIP. ITEM")] for row in rows[1:]
        ]

        assert descripciones == ["Segundo item"]
